#!/usr/bin/env python3
"""
网络质差整治通报 - 一键生成脚本（纯 openpyxl，无截图）
功能：
  1. 将工单清单数据复制到模板 → 生成 Excel 通报
  2. 输出到 Desktop/test/output/ 目录

使用方式：
  - 命令行：python auto_report.py
  - 双击：   run_auto_report.bat
"""

import openpyxl
import os
import datetime
import time

# ── 路径配置 ──────────────────────────────────────────────────────────────────
GD_FILE   = r'C:\Users\zml95\Desktop\test\网络质差整治工单清单.xlsx'
TPL_FILE  = r'C:\Users\zml95\Desktop\test\主动服务-网络质差整治通报版本.xlsx'
OUTPUT_DIR = r'C:\Users\zml95\Desktop\test\output'

now       = datetime.datetime.now()
timestamp = now.strftime('%Y%m%d_%H%M%S')


def step1_generate_excel():
    """
    Step1：用 openpyxl 的 read_only 模式读源文件，
           再写入模板"质差整治清单"sheet 的 H 列开始位置。
    """
    print('=' * 60)
    print('Step 1: 生成 Excel 通报文件...')
    t0 = time.time()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── 1. 以只读模式快速读取源文件 ────────────────────────────────
    print('  ① 读取工单清单（只读模式）...', flush=True)
    wb_src = openpyxl.load_workbook(GD_FILE, read_only=True, data_only=True)
    ws_src = wb_src.active

    # 用 values 属性一次性拿到所有数据（返回 generator）
    all_rows = list(ws_src.values)   # list of tuples
    wb_src.close()

    src_rows = len(all_rows)      # 含表头
    src_cols = max((len(r) for r in all_rows), default=0)
    data_rows = src_rows - 1      # 去掉表头
    print(f'    共 {src_rows} 行（含表头 {src_cols} 列），'
          f'数据 {data_rows} 行', flush=True)

    # ── 2. 打开模板，准备目标 sheet ───────────────────────────────────
    print('  ② 打开模板...', flush=True)
    wb_dst = openpyxl.load_workbook(TPL_FILE)
    ws_dst = wb_dst['质差整治清单']

    # 清空 H 列及之后第 2 行起的旧数据（只清空有数据的区域）
    old_max_row = ws_dst.max_row
    if old_max_row > 1:
        print(f'  ③ 清空旧数据（{old_max_row - 1} 行）...', flush=True)
        for row in range(2, old_max_row + 1):
            for col in range(8, ws_dst.max_column + 1):
                ws_dst.cell(row=row, column=col).value = None

    # ── 3. 批量写入新数据（H 列 = 第 8 列）────────────────────
    print(f'  ④ 写入新数据（{data_rows} 行 × {src_cols} 列）...')
    # 写表头（第 1 行）
    header = all_rows[0]
    for ci, val in enumerate(header, start=8):
        ws_dst.cell(row=1, column=ci).value = val

    # 写数据行（第 2 行开始）
    total_cells = data_rows * min(src_cols, 300)   # 只写前 300 列
    written = 0
    report_every = max(total_cells // 20, 1)

    for ri, row_data in enumerate(all_rows[1:], start=2):
        for ci, val in enumerate(row_data, start=8):
            if ci - 8 >= 300:          # 最多写前 300 列（足够覆盖关键列）
                break
            if val is not None:
                ws_dst.cell(row=ri, column=ci).value = val
            written += 1
        if written % report_every < (ci - 8 if 'ci' in dir() else 300):
            pct = written * 100 // total_cells
            print(f'    ...{pct}%', end='\r', flush=True)
    print(f'    ...100%{" "*10}')

    # ── 4. 向下填充 A-G 列公式 ─────────────────────────────────────
    print('  ⑤ 填充 A-G 列公式...', flush=True)
    # 收集第 2 行的公式
    formula_tpl = {}
    for col_letter in ['A','B','C','D','E','F','G']:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws_dst.cell(row=2, column=col_idx)
        if cell.value and str(cell.value).startswith('='):
            formula_tpl[col_letter] = str(cell.value)

    # 为第 3～N 行生成公式（替换行号为当前行）
    import re
    row_refs = set()
    for ft in formula_tpl.values():
        for m in re.finditer(r'([A-Z]+)2\b', ft):
            row_refs.add(m.group(1))
    row_refs = sorted(row_refs)   # e.g. ['A','AK','B','BS','DB','FN'...]

    for ri in range(3, data_rows + 2):
        for col_letter, base in formula_tpl.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)
            new_f = base
            for ref in row_refs:
                new_f = new_f.replace(f'{ref}2', f'{ref}{ri}')
            ws_dst.cell(row=ri, column=col_idx).value = new_f
    print(f'    公式已填充到第 {data_rows + 1} 行', flush=True)

    # ── 5. 保存 ─────────────────────────────────────────────────────
    excel_path = os.path.join(OUTPUT_DIR, f'网络质差整治通报_{timestamp}.xlsx')
    print(f'  ⑥ 保存文件 → {excel_path} ...', flush=True)
    wb_dst.save(excel_path)
    wb_dst.close()

    elapsed = time.time() - t0
    print(f'  ✅ Excel 生成完毕！用时 {elapsed:.1f} 秒')
    print(f'     文件：{excel_path}')
    print(f'     数据：{data_rows} 行 × {min(src_cols, 300)} 列')
    return excel_path


def main():
    print()
    print('╔══════════════════════════════════════════╗')
    print('║   网络质差整治通报 — 一键生成（Excel版） ║')
    print('╚══════════════════════════════════════════╝')
    print(f'  时间：{now.strftime("%Y-%m-%d %H:%M:%S")}')
    print()

    t0 = time.time()
    excel_path = step1_generate_excel()

    elapsed = time.time() - t0
    print()
    print('=' * 60)
    print(f'✅ 全部完成！总用时 {elapsed:.1f} 秒')
    print(f'  Excel 文件：{excel_path}')
    print()
    print('  后续步骤：')
    print('  1. 用 WPS/Excel 打开上面的 Excel 文件')
    print('  2. "通报" sheet 的数据透视和公式会自动刷新')
    print('  3. 截图发到微信群')
    print('=' * 60)


if __name__ == '__main__':
    main()
